#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2020/3/12 22:03
# @Author  : Roohom
# @Site    : 
# @File    : excel2.py
# @Software: PyCharm
"""
读取excel文件
"""

from openpyxl import load_workbook
from openpyxl import Workbook

workbook = load_workbook('王者峡谷数据.xlsx')
print(workbook.sheetnames)
sheet = workbook[workbook.sheetnames[0]]
print(sheet.title)

for row in range(2, 7):
    for col in range(65, 70):
        cell_index = chr(col) + str(row)
        print(sheet[cell_index].value, end='\t')
    print()
